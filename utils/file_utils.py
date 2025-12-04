import pandas as pd
import warnings
import os
from openpyxl import load_workbook
from typing import Tuple, Dict, Any

import os
import pandas as pd
import warnings
from openpyxl import load_workbook
from typing import Tuple, Dict, Any

def validate_excel_structure(file_path: str) -> Tuple[bool, str]:
    """验证Excel文件结构并进行表头映射"""
    try:
        # 直接使用pandas读取，让它自动选择引擎
        try:
            df = pd.read_excel(file_path, engine=None, nrows=1)
        except Exception as e:
            # 如果自动选择失败，尝试手动选择
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext == '.xls':
                # 检查是否安装了xlrd
                try:
                    import xlrd
                    df = pd.read_excel(file_path, engine='xlrd', nrows=1)
                except ImportError:
                    return False, "需要安装xlrd库来处理.xls文件。请运行: pip install xlrd"
            else:
                # 尝试openpyxl
                df = pd.read_excel(file_path, engine='openpyxl', nrows=1)
        
        # 获取表头
        headers = [str(col).strip() for col in df.columns]
        
        # 应用表头映射
        mapped_headers = []
        header_mapping = {
            '备注（小客户名称）': '子客户名称',
            '票 号': '票号',
            '品牌': '备注'
        }
        
        for header in headers:
            header_no_space = header.replace(' ', '')
            mapped_header = header
            for original, standard in header_mapping.items():
                if header_no_space == original.replace(' ', ''):
                    mapped_header = standard
                    break
            mapped_headers.append(mapped_header)
        
        # 必需的表头（映射后的标准表头）
        required_headers = ['客户名称', '编号', '子客户名称', '年', '月', '日', '收款金额', '颜色', '等级', '数量', '单价', '金额', '余额', '票号', '备注', '生产线']
        
        # 去除空格后的表头用于匹配
        mapped_headers_no_space = [h.replace(' ', '') for h in mapped_headers]
        required_headers_no_space = [h.replace(' ', '') for h in required_headers]
        
        missing_headers = []
        for req_header, req_header_ns in zip(required_headers, required_headers_no_space):
            if req_header_ns not in mapped_headers_no_space:
                missing_headers.append(req_header)
        
        if missing_headers:
            return False, f"缺少必要的表头: {missing_headers}"
        
        return True, "文件结构正确"
        
    except Exception as e:
        return False, f"文件检查失败: {str(e)}"

def preview_excel_data(file_path: str, nrows: int = 5) -> Tuple[bool, pd.DataFrame]:
    """预览Excel数据（应用表头映射），支持.xls和.xlsx格式"""
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            
            # 根据文件扩展名选择引擎
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext == '.xls':
                engine = 'xlrd'
            else:
                engine = 'openpyxl'
            
            # 先读取原始数据
            df = pd.read_excel(file_path, engine=engine, nrows=nrows)
            
            # 应用表头映射
            df = apply_header_mapping(df)
            
        return True, df
    except Exception as e:
        return False, f"无法读取文件: {str(e)}"

def apply_header_mapping(df: pd.DataFrame) -> pd.DataFrame:
    """应用表头映射到DataFrame"""
    # 定义表头映射关系
    header_mapping = {
        '备注（小客户名称）': '子客户名称',
        '票 号': '票号',
        '品牌': '备注'
    }
    
    # 创建新的列名列表
    new_columns = []
    for col in df.columns:
        original_col = str(col).strip()
        # 去除空格进行匹配
        original_col_no_space = original_col.replace(' ', '')
        mapped_col = original_col
        
        # 检查映射关系
        for original, standard in header_mapping.items():
            if original_col_no_space == original.replace(' ', ''):
                mapped_col = standard
                break
        
        new_columns.append(mapped_col)
    
    # 应用新的列名
    df_mapped = df.copy()
    df_mapped.columns = new_columns
    
    return df_mapped

def get_excel_file_info(file_path: str) -> Dict[str, Any]:
    """获取Excel文件的详细信息（应用表头映射），支持.xls和.xlsx格式"""
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.xls':
            # 使用xlrd读取.xls文件
            import xlrd
            wb = xlrd.open_workbook(file_path, on_demand=True)
            sheet = wb.sheet_by_index(0)
            
            # 获取原始表头
            original_headers = []
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(0, col_idx)
                if cell_value is not None:
                    original_headers.append(str(cell_value).strip())
            
            # 统计行数（不包括表头）
            row_count = 0
            for row_idx in range(1, min(100000, sheet.nrows)):
                if sheet.cell_value(row_idx, 0) is not None:
                    row_count += 1
            
            wb.release_resources()
            
        elif file_ext in ['.xlsx', '.xlsm']:
            # 使用openpyxl读取.xlsx文件
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                wb = load_workbook(file_path, data_only=True, read_only=True)
                sheet = wb.active
                
                # 获取原始表头
                original_headers = []
                for cell in sheet[1]:
                    if cell.value is not None:
                        original_headers.append(str(cell.value).strip())
                
                # 统计行数（不包括表头）
                row_count = 0
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=100000), 2):
                    if row[0].value is not None:
                        row_count += 1
                    else:
                        break
                
                wb.close()
        else:
            return {"error": f"不支持的文件格式: {file_ext}"}
        
        # 应用表头映射
        mapped_headers = []
        header_mapping = {
            '备注（小客户名称）': '子客户名称',
            '票 号': '票号',
            '品牌': '备注'
        }
        
        for header in original_headers:
            header_no_space = header.replace(' ', '')
            mapped_header = header
            for original, standard in header_mapping.items():
                if header_no_space == original.replace(' ', ''):
                    mapped_header = standard
                    break
            mapped_headers.append(mapped_header)
        
        # 读取数据并应用表头映射
        engine = 'xlrd' if file_ext == '.xls' else 'openpyxl'
        df = pd.read_excel(file_path, engine=engine, nrows=1000)
        df = apply_header_mapping(df)
        
        # 创建去除空格后的列名映射
        column_mapping = {}
        for col in df.columns:
            col_no_space = col.replace(' ', '')
            column_mapping[col_no_space] = col
        
        info = {
            "original_headers": original_headers,
            "mapped_headers": mapped_headers,
            "headers": mapped_headers,  # 保持兼容性
            "row_count": row_count,
            "column_count": len(mapped_headers),
            "required_headers_present": True,
            "sample_data_available": len(df) > 0 if not df.empty else False,
            "column_mapping": column_mapping
        }
        
        # 如果数据不为空，添加更多统计信息
        if not df.empty:
            # 使用映射后的列名
            customer_col = '客户名称'
            product_col = '产品名称' if '产品名称' in df.columns else None
            color_col = '颜色'
            quantity_col = '数量'
            price_col = '单价'
            amount_col = '金额'
            
            info.update({
                "customer_count": df[customer_col].nunique() if customer_col in df.columns else 0,
                "product_count": df[product_col].nunique() if product_col and product_col in df.columns else 0,
                "color_count": df[color_col].nunique() if color_col in df.columns else 0,
                "has_numeric_data": any(col in df.columns for col in [quantity_col, price_col, amount_col])
            })
        
        return info
        
    except Exception as e:
        return {"error": f"获取文件信息失败: {str(e)}"}

def get_import_strategy_description(strategy: str) -> Dict[str, Any]:
    """获取导入策略的详细说明"""
    descriptions = {
        "update": {
            "name": "智能更新",
            "description": "更新重复数据，添加新数据，保持数据同步",
            "icon": "📝",
            "recommended": True,
            "details": [
                "✅ 更新重复的销售记录",
                "✅ 添加新的销售记录",  
                "✅ 保持客户信息同步更新",
                "💡 **适用场景**: 日常数据更新、修正错误数据"
            ]
        },
        "append": {
            "name": "仅新增", 
            "description": "只导入不存在的新数据，不修改已有记录",
            "icon": "➕",
            "recommended": False,
            "details": [
                "✅ 只导入不存在的新数据",
                "❌ 不修改任何已有记录", 
                "💡 **适用场景**: 补充历史数据、避免数据冲突"
            ]
        },
        "replace": {
            "name": "完全替换",
            "description": "清空所有数据后重新导入完整数据集", 
            "icon": "🔄",
            "recommended": False,
            "details": [
                "⚠️ 清空所有现有数据",
                "⚠️ 重新导入完整数据集",
                "💡 **适用场景**: 数据重构、重大结构调整"
            ]
        }
    }
    return descriptions.get(strategy, descriptions["update"])

def validate_data_quality(df: pd.DataFrame) -> Dict[str, Any]:
    """验证数据质量（使用映射后的列名）"""
    issues = []
    warnings = []
    
    # 应用表头映射
    df_mapped = apply_header_mapping(df)
    
    # 创建去除空格后的列名映射
    column_mapping = {}
    for col in df_mapped.columns:
        col_no_space = col.replace(' ', '')
        column_mapping[col_no_space] = col
    
    # 检查必需字段的空值
    required_fields = ['客户名称', '编号']
    for field in required_fields:
        actual_col = column_mapping.get(field, field)
        if actual_col in df_mapped.columns:
            null_count = df_mapped[actual_col].isnull().sum()
            if null_count > 0:
                issues.append(f"字段 '{actual_col}' 有 {null_count} 个空值")
    
    # 检查数值字段的有效性
    numeric_fields = ['数量', '单价', '金额', '收款金额']
    for field in numeric_fields:
        actual_col = column_mapping.get(field, field)
        if actual_col in df_mapped.columns:
            # 检查异常大值（假设大于100万为异常）
            large_value_count = (df_mapped[actual_col] > 1000000).sum()
            if large_value_count > 0:
                warnings.append(f"字段 '{actual_col}' 有 {large_value_count} 个异常大值")
    
    # 检查日期字段的有效性
    date_fields = ['年', '月', '日']
    for field in date_fields:
        actual_col = column_mapping.get(field, field)
        if actual_col in df_mapped.columns:
            invalid_count = df_mapped[actual_col].isnull().sum()
            if invalid_count > 0:
                warnings.append(f"字段 '{actual_col}' 有 {invalid_count} 个无效值")
    
    return {
        "has_issues": len(issues) > 0,
        "has_warnings": len(warnings) > 0,
        "issues": issues,
        "warnings": warnings,
        "total_records": len(df_mapped),
        "valid_records": len(df_mapped) - len(issues)
    }

def get_recommended_strategy(file_info: Dict[str, Any], db_status: Dict[str, Any]) -> str:
    """根据文件信息和数据库状态推荐导入策略"""
    
    # 如果数据库为空，推荐替换模式
    if db_status.get('sales_records_count', 0) == 0:
        return "replace"
    
    # 如果文件包含大量数据且数据库已有数据，推荐更新模式
    if file_info.get('row_count', 0) > 1000 and db_status.get('sales_records_count', 0) > 0:
        return "update"
    
    # 如果文件数据量较小，推荐追加模式
    if file_info.get('row_count', 0) < 100:
        return "append"
    
    # 默认推荐更新模式
    return "update"

# 新增函数：标准化列名（去除空格）
def standardize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """标准化DataFrame的列名，去除空格"""
    df_standardized = df.copy()
    df_standardized.columns = [col.replace(' ', '') for col in df_standardized.columns]
    return df_standardized

# 新增函数：获取原始列名到标准列名的映射
def get_column_mapping(headers: list) -> Dict[str, str]:
    """获取原始列名到标准列名（去除空格）的映射"""
    mapping = {}
    for header in headers:
        standard_header = header.replace(' ', '')
        mapping[standard_header] = header
    return mapping